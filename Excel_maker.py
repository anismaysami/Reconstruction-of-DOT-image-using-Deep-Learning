# -*- coding: utf-8 -*-
"""
Created on Tue Dec 20 09:55:40 2022

@author: Anis Maysami
"""
#This function write a data frame to an .xlsx file

#Importing needed package
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#This function read python data to a pandas data frame and
# write a pandas data frame to xlsx file
def create_xlsx(data, header, workbook_name, sheet_name):
    """
    data is a numpy array or a list
    header is string list
    workbook_name is a string that is excel file name
    sheet_name is a string that is sheet name of exel file
    """
    wb = Workbook()
    page = wb.active

    page.title =sheet_name #sheet title
    page.append(header) # write the headers to the first line

    for info in data:
        page.append(info)
    wb.save(filename = workbook_name)

def rewrite_xlsx(workbook_name, new_data):

    wb = load_workbook(workbook_name)
    page = wb.active

    # New data to write:
    for info in new_data:
        page.append(info)
    wb.save(filename=workbook_name)

if __name__ == "__main__":
    data=[[1,1,1]]
    header=['1', '2','3']
    workbook_name='optoskin.xlsx'
    sheet_name='A'
    create_xlsx(data, header, workbook_name, sheet_name)
    rewrite_xlsx(workbook_name, data)